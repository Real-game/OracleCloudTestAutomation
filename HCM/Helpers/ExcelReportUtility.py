import os
import xlrd
import pandas as pd
from robot.api.deco import keyword
import getpass
from openpyxl import load_workbook
import openpyxl

@keyword
def get_column_data(column_name, file_name, sheet_index=0):
    expected_column_index = 0
    expected_row_index = 0
    excel_all_data = xlrd.open_workbook(file_name)
    excelsheet0_data = excel_all_data.sheet_by_index(sheet_index)
    for row in range(excelsheet0_data.nrows):
        for column in range(excelsheet0_data.ncols):
            if excelsheet0_data.cell_value(row, column) == column_name:
                expected_column_index = column
                expected_row_index = row
                break
    column_data = excelsheet0_data.col_values(expected_column_index, expected_row_index + 1, excelsheet0_data.nrows)
    return str(column_data)


@keyword
def compare_excel_all_columns(excel_1, excel_2,excel_1_start_row=0, excel_2_start_row=0):
    status_flag = "True"
    if excel_1_start_row != 0: excel_1_start_row = excel_1_start_row - 1
    if excel_2_start_row != 0: excel_2_start_row = excel_2_start_row - 1
    excel_data_1 = pd.read_excel(os.getcwd() + "/TestData/" + excel_1+".xls", skiprows = excel_1_start_row)
    excel_data_2 = pd.read_excel(os.path.expanduser("~") + "/Downloads/" + excel_2+".xls", skiprows = excel_2_start_row)
    excel_1_column_name = list(excel_data_1.columns)
    excel_2_column_name = list(excel_data_2.columns)
    if len(excel_1_column_name) == len(excel_2_column_name):
        for column_name in excel_1_column_name:
                print("Checking column : " + str(column_name))
                excel_1_column_data = get_column_data(column_name, os.getcwd() + "/TestData/" + excel_1 + ".xls")
                excel_2_column_data = get_column_data(column_name,
                                                      os.path.expanduser("~") + "/Downloads/" + excel_2 + ".xls")
                if excel_1_column_data != excel_2_column_data:
                    status_flag = "False"
                    x = excel_1_column_data.split(",")
                    y = excel_2_column_data.split(",")
                    for i in range(len(x)):
                        if x[i] != y[i]:
                            print(x[i])
                            print("Mismatch value: Before upgrade in the column " + column_name + " value is " + x[
                                i] + " and after upgrade the value is changed to " + y[i])
                else:
                    print("Matched")
    else:
        print("Column count mismatched")
        status_flag = "False"
    return status_flag


@keyword
def compare_excel_skip_columns(excel_1, excel_2, column_list, excel_1_start_row=0, excel_2_start_row=0):
    status_flag = "True"
    if excel_1_start_row != 0: excel_1_start_row = excel_1_start_row - 1
    if excel_2_start_row != 0: excel_2_start_row = excel_2_start_row - 1
    excel_data_1 = pd.read_excel(os.getcwd() + "/TestData/" + excel_1+".xls", skiprows = excel_1_start_row)
    excel_data_2 = pd.read_excel(os.path.expanduser("~") + "/Downloads/" + excel_2+".xls", skiprows = excel_2_start_row)
    excel_1_column_name = list(excel_data_1.columns)
    excel_2_column_name = list(excel_data_2.columns)
    if len(excel_1_column_name) == len(excel_2_column_name):
        for column_name in excel_1_column_name:
            if column_name not in column_list:
                print("Checking column : "+str(column_name))
                excel_1_column_data = get_column_data(column_name, os.getcwd() + "/TestData/" + excel_1+".xls")
                excel_2_column_data = get_column_data(column_name, os.path.expanduser("~") + "/Downloads/" + excel_2+".xls")
                if excel_1_column_data != excel_2_column_data:
                    status_flag = "False"
                    x=excel_1_column_data.split(",")
                    y=excel_2_column_data.split(",")
                    for i in range(len(x)):
                        if x[i] != y[i]:
                            print(x[i])
                            print("Mismatch value: Before upgrade in the column "+column_name+" value is "+x[i]+" and after upgrade the value is changed to "+y[i])
                else:
                    print("Matched")
            else:
                print("Skipping Column : "+column_name)
    else:
        print("Column count mismatched")
        status_flag = "False"
    return status_flag


@keyword
def compare_excel_specific_columns(excel_1, excel_2, column_list, excel_1_start_row=0, excel_2_start_row=0):
    status_flag = "True"
    if excel_1_start_row != 0: excel_1_start_row = excel_1_start_row - 1
    if excel_2_start_row != 0: excel_2_start_row = excel_2_start_row - 1
    excel_data_1 = pd.read_excel(os.getcwd() + "/TestData/" + excel_1+".xls", skiprows = excel_1_start_row)
    excel_data_2 = pd.read_excel(os.path.expanduser("~") + "/Downloads/" + excel_2+".xls", skiprows = excel_2_start_row)
    excel_1_column_name = list(excel_data_1.columns)
    excel_2_column_name = list(excel_data_2.columns)
    if len(excel_1_column_name) == len(excel_2_column_name):
        for column_name in excel_1_column_name:
            if column_name in column_list:
                print("Checking column : "+str(column_name))
                excel_1_column_data = get_column_data(column_name, os.getcwd() + "/TestData/" + excel_1+".xls")
                excel_2_column_data = get_column_data(column_name, os.path.expanduser("~") + "/Downloads/" + excel_2+".xls")
                if excel_1_column_data != excel_2_column_data:
                    status_flag = "False"
                    x=excel_1_column_data.split(",")
                    y=excel_2_column_data.split(",")
                    for i in range(len(x)):
                        if x[i] != y[i]:
                            print(x[i])
                            print("Mismatch value: Before upgrade in the column "+column_name+" value is "+x[i]+" and after upgrade the value is changed to "+y[i])

                else:
                    print("Matched")
            else:
                print("Skipping Column : "+column_name)
    else:
        print("Column count mismatched")
        status_flag = "False"
    return status_flag


# @keyword
# def delete_File(file_name):
#     default_download_path = os.path.expanduser("~") + "/Downloads/" + file_name + ".xls"
#
#     if os.path.exists(default_download_path):
#         os.remove(default_download_path)
#     else:
#         print("The file does not exist")
#         #return "Absent"


# @keyword
# def get_Employee_Number_Search_Validation(file_name, first_col_header, required_col_header, validation_value):
#     user_name = getpass.getuser()
#     # default_download_path = r'C:\\Users\\' + user_name + '\\Downloads\\' + file_name + '.xls'  # windows
#     default_download_path = os.path.expanduser("~") + "/Downloads/" + file_name + ".xls"
#     py_xlrd = xlrd.open_workbook(default_download_path)
#     py_sheet = py_xlrd.sheet_by_index(0)
#     employee_col = -1
#     employee_row = -1
#     found = 0
#
#     for i in range(py_sheet.nrows):
#         if py_sheet.cell_value(i, 0).strip() == first_col_header:
#             header_row = i
#             break
#
#     for i in range(py_sheet.ncols):
#         if py_sheet.cell_value(header_row, i).strip() == required_col_header:
#             employee_col = i
#             break
#
#     for i in range(header_row + 1, py_sheet.nrows):
#         # py_sheet.cell_value(i, employee_col).strip()
#         if py_sheet.cell_value(i, employee_col).strip() == validation_value:
#             found = 1
#             break
#
#     return found

# @keyword
# def compare_my_excel_files(excel_1, excel_2, skip_rows):
#     status_flag = "True"
#     # Read the Excel files into DataFrames, skipping the specified number of rows
#     df1 = pd.read_excel(os.getcwd() + "/TestData/" + excel_1+".xlsx", skiprows=skip_rows)
#     df2 = pd.read_excel(os.path.expanduser("~") + "/Downloads/" + excel_2+".xlsx", skiprows=skip_rows)
#
#     # Check if the DataFrames are equal
#     are_equal = df1.equals(df2)
#
#     if are_equal:
#         print("The Excel files are identical.")
#     else:
#         # Find the mismatched cells by comparing the DataFrames
#         mismatched_cells = (df1 != df2)
#         status_flag = "False"
#         # Filter the DataFrame to show only the mismatched cells
#         mismatched_data = df1.where(mismatched_cells)
#
#         # Print the mismatched data
#         print("Mismatched data:")
#         print(mismatched_data)
#     return status_flag

# @keyword
# def compare_excel_mine_all_columns(excel_1, excel_2, excel_1_start_row=0, excel_2_start_row=0):
#     status_flag = True
#     # if excel_1_start_row != 0:
#     #     excel_1_start_row = excel_1_start_row - 1
#     #     print(excel_1_start_row)
#     # if excel_2_start_row != 0:
#     #     excel_2_start_row = excel_2_start_row - 1
#     # print(excel_2_start_row)
#     excel_data_1 = load_workbook(os.getcwd() + "/TestData/" + excel_1 + ".xlsx")
#     sheet_1 = excel_data_1.active
#     print(os.getcwd() + "/TestData/" + excel_1 + ".xlsx")
#     print(excel_data_1)
#     print(sheet_1)
#     excel_data_2 = load_workbook(os.path.expanduser("~") + "/Downloads/" + excel_2 + ".xlsx")
#     sheet_2 = excel_data_2.active
#     print(os.path.expanduser("~") + "/Downloads/" + excel_2 + ".xlsx")
#     print(sheet_2)
#     print(excel_data_2)
#     excel_1_column_name = [cell.value for cell in sheet_1[excel_1_start_row]]
#     print(excel_1_column_name)
#     excel_2_column_name = [cell.value for cell in sheet_2[excel_2_start_row]]
#     print(excel_2_column_name)
#     if len(excel_1_column_name) == len(excel_2_column_name):
#         print(len(excel_1_column_name))
#         print(len(excel_2_column_name))
#         for column_name in excel_1_column_name:
#             print("Checking column: " + str(column_name))
#             excel_1_column_data = get_column_mydata(column_name, os.getcwd() + "/TestData/" + excel_1 + ".xlsx")
#             print("7")
#             excel_2_column_data = get_column_mydata(column_name, os.path.expanduser("~") + "/Downloads/" + excel_2 + ".xlsx")
#             print("8")
#             if excel_1_column_data != excel_2_column_data:
#                 status_flag = False
#                 print(status_flag)
#                 x = excel_1_column_data.split(",")
#                 y = excel_2_column_data.split(",")
#                 for i in range(len(x)):
#                     if x[i] != y[i]:
#                         print(x[i])
#                         print("Mismatch value: Before upgrade in the column " + column_name + " value is " + x[i] + " and after upgrade the value is changed to " + y[i])
#             else:
#                 print("Matched")
#     else:
#         print("Column count mismatched")
#         status_flag = False
#     return status_flag

# @keyword
# def get_mycolumn_data(column_name, file_name, sheet_name=None):
#     column_data = []
#     try:
#         workbook = load_workbook(file_name)
#         sheet = workbook.active if sheet_name is None else workbook[sheet_name]
#         # Find the column index for the specified column name
#         expected_column_index = None
#         for col_idx, cell_value in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)), start=1):
#             if cell_value == column_name:
#                 expected_column_index = col_idx
#                 break
#         if expected_column_index:
#             # Iterate through the rows and extract data from the specified column
#             for row in sheet.iter_rows(min_row=2, values_only=True):
#                 cell_value = row[expected_column_index - 1]
#                 if cell_value is not None:
#                     column_data.append(str(cell_value))
#         else:
#             print(f"Column '{column_name}' not found in the Excel file.")
#     except FileNotFoundError:
#         print(f"File not found: '{file_name}'.")
#     return ','.join(column_data)


# @keyword
# def get_column_mydata(column_name, file_name, sheet_index=0):
#     expected_column_index = 0
#     expected_row_index = 0
#     workbook = openpyxl.load_workbook(file_name)
#     print(file_name)
#     sheet = workbook.worksheets[sheet_index]
#     print(sheet_index)
#     print(sheet)
#     print(sheet.iter_rows())
#     print("1")
#     for row in sheet.iter_rows():
#         print(sheet.iter_rows())
#         for cell in row:
#             print(cell.value)
#             if cell.value == column_name:
#                 expected_column_index = cell.column - 1
#                 print(expected_column_index,cell.column - 1)
#                 expected_row_index = cell.row - 1
#                 print(expected_row_index,cell.row - 1)
#                 break
#     column_data = [cell.value for cell in sheet[f"{column_name}!{expected_row_index+1}:{column_name}!{sheet.iter_rows()}"]]
#     return str(column_data)
