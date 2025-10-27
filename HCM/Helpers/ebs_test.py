import string
import pandas as pd
import openpyxl
import re
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")



def extract_internal_note(file2_path,sheetname):
    df = pd.read_excel("C:\\Users\\senthil.n.thangaraj\\Downloads/" + file2_path + ".xlsx", sheet_name=sheetname)
    # cell_value = df.iloc[2][3]  # Accenture dataset excel
    cell_value = df.iloc[0, 3]  # TestFile RFP Entry
    # cell_value = df.iloc[0, 8]   #RFP-2024-ITIN-512 - Technical Requirements Evaluator Scoring - All Consensus - 2024-10-21
    return cell_value

def extract_internal_note_consensus(file2_path,sheetname):
    df = pd.read_excel("C:\\Users\\senthil.n.thangaraj\\Downloads/" + file2_path + ".xlsx", sheet_name=sheetname)
    # cell_value = df.iloc[2][3]  # Accenture dataset excel
    # cell_value = df.iloc[0, 3]  # TestFile RFP Entry
    cell_value = df.iloc[0, 8]   #RFP-2024-ITIN-512 - Technical Requirements Evaluator Scoring - All Consensus - 2024-10-21
    return cell_value

def normalize_text(text):
    # Strip leading/trailing whitespace, convert to lowercase, remove punctuation, and normalize spaces
    if text:
        # Remove punctuation, convert to lowercase, and replace multiple spaces with a single space
        text = text.strip().lower().translate(str.maketrans('', '', string.punctuation))
        text = re.sub(r'\s+', ' ', text)  # Replace multiple spaces with a single space
    return text

def get_question_details_from_excel(file_path, question,sheetname):
    wb = openpyxl.load_workbook("C:\\Users\\senthil.n.thangaraj\\Downloads/" + file_path + ".xlsx")
    sheet = wb[sheetname]  # Replace 'EntrySheet' with your actual sheet name
    normalized_question = normalize_text(question)  # Normalize the input question
    for row in sheet.iter_rows(min_row=3):  # Start from row 3
        excel_question = row[3].value  # D is the 4th column (index 3)
        # if excel_question == question:
        if normalize_text(excel_question) == normalized_question:  # Normalize comparison
            score = row[7].value  # H is the 8th column (index 7)
            # print(type(score))
            note = row[8].value   # I is the 9th column (index 8)
            # print(type(note))
            return score, note
    return None, None  # If question not found

def validate_score(value):
    try:
        # Convert value to an integer
        score = int(value)
        # Check if the score is greater than 10
        if score > 10:
            raise ValueError(f"Warning: Score is greater than 10: {score}")
        return score
    except ValueError:
        # If conversion to integer fails or any other error occurs, raise a failure
        raise ValueError(f"Score is not a valid integer or not within the acceptable range: {value}")

def get_column_values(filepath, sheet_name, column='D', start_row=3):
    workbook = openpyxl.load_workbook("C:\\Users\\senthil.n.thangaraj\\Downloads/" + filepath + ".xlsx")
    sheet = workbook[sheet_name]
    values = []
    for row in range(start_row, sheet.max_row + 1):
        cell_value = sheet[f"{column}{row}"].value
        if cell_value is not None:
            values.append(cell_value)
    return values

# def get_evaluator_question_details_from_excel(file_path, question, sheetname, evaluator_name):
#     # Suppress the UserWarning
#     warnings.simplefilter("ignore", category=UserWarning)
#     # Load the workbook and select the sheet
#     wb = openpyxl.load_workbook("C:\\Users\\senthil.n.thangaraj\\Downloads/" + file_path + ".xlsx")
#     sheet = wb[sheetname]
#
#     # Normalize the input question
#     normalized_question = normalize_text(question)
#
#     # Dictionary to store evaluator details
#     evaluator_details = {}
#     active_question = None  # To track the currently active question
#
#     # Iterate through rows to find the matching criteria
#     for row in sheet.iter_rows(min_row=7):
#         # Column D (index 3) contains Criteria
#         excel_question = row[3].value if row[3].value else None
#
#         # Update the active question if a new question is found
#         if excel_question:
#             active_question = normalize_text(excel_question)
#
#         # If the current row belongs to the matching question
#         if active_question == normalized_question:
#             # Iterate through evaluator details
#             for col in range(7, len(row), 3):  # Start at Column H (index 7) and step by 3
#                 eval_name = row[col].value  # Column H (Evaluator Name)
#
#                 # Check if the evaluator name is valid (a string) and exclude formulas
#                 if eval_name and isinstance(eval_name, str) and not eval_name.startswith("="):
#                     note = row[col + 1].value if col + 1 < len(row) else None  # Column I (Note)
#                     # score = row[col + 2].value if col + 2 < len(row) else None  # Column J (Score)
#                     score = row[col + 2].value if col + 2 < len(row) and row[col + 2].value is not None else "N/A"
#
#                     # Add the details to the dictionary
#                     evaluator_details[eval_name] = {
#                         "note": note if note else "N/A",  # Handle missing notes
#                         # "score": score if score else "N/A"  # Handle missing scores
#                         "score": score  # Keep 0 as a valid score
#                     }
#
#     if evaluator_name in evaluator_details:
#         return evaluator_details[evaluator_name]['score'], evaluator_details[evaluator_name]['note']
#     else:
#         return "N/A", "N/A"  # If the evaluator is not found

def get_evaluator_question_details_from_excel(file_path, question, sheetname, evaluator_name):
    # Load the Excel sheet into a pandas DataFrame
    file_path_full = f"C:\\Users\\senthil.n.thangaraj\\Downloads\\{file_path}.xlsx"
    df = pd.read_excel(file_path_full, sheet_name=sheetname, header=None)

    # Normalize the input question
    normalized_question = normalize_text(question)

    # Initialize evaluator details
    evaluator_details = {}
    active_question = None

    # Iterate through DataFrame rows
    for _, row in df.iterrows():
        # Column D (index 3) contains the question
        excel_question = row[3] if not pd.isna(row[3]) else None

        # Update the active question if a new one is found
        if excel_question:
            active_question = normalize_text(excel_question)

        # If the current row belongs to the matching question
        if active_question == normalized_question:
            # Iterate through evaluator details starting at Column H (index 7)
            for col_index in range(7, len(row), 3):  # Start at Column H and step by 3
                eval_name = row[col_index]

                # Check if the evaluator name is valid and not NaN
                if pd.notna(eval_name) and isinstance(eval_name, str):
                    note = row[col_index + 1] if col_index + 1 < len(row) and pd.notna(row[col_index + 1]) else "N/A"
                    score = row[col_index + 2] if col_index + 2 < len(row) and pd.notna(row[col_index + 2]) else "N/A"

                    # Store details in the dictionary
                    evaluator_details[eval_name] = {
                        "note": note,
                        "score": score,
                    }

    # Return the evaluator's details if available
    if evaluator_name in evaluator_details:
        return evaluator_details[evaluator_name]["score"], evaluator_details[evaluator_name]["note"]
    else:
        return "N/A", "N/A"

if __name__ == "__main__":
    file1 = "RFP-2024-ITIN-512 - Technical Requirements Evaluator Scoring - All Consensus - 2024-10-21"
    sheetname1 = "Pandos"
    evaluator= "Marise"

    file = "TestFile RFP Entry"
    sheetname = "EntrySheet"
    question = "Provides APIs for integration (Push, Pull, Sync) allow import and export of data e.g. to Records Management, Data Warehouse and Azure Data Lake with immediate reporting data structures."
    details = get_evaluator_question_details_from_excel(file1, question,sheetname1,evaluator)

    # details = get_column_values(file1, sheetname)
    print(details)
    DET = get_question_details_from_excel(file,question,sheetname)
    # NOTE = extract_internal_note(file,sheetname)
    print(DET)

# if __name__ == "__main__":
#     file1 = "TestFile RFP Entry"
#     # file1 = "Accenture Dataset"
#     question = "The platform shall allow Metrolinx to import flat files (csv, xlsx, etc.) on an automated and scheduled basis as well as on a manual and ad-hoc basis. The refresh frequency of scheduled import is daily.\n\nPlease describe how your solution meets these requirements."
#     details = extract_internal_note(file1)
#     print(details)
# Example test
# test_value = 2
# result = validate_score(test_value)
# print(result)

